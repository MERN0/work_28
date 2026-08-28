"""Writes the SYS5 output workbook: Cover Page, Test Pattern, Test Cases,
Item List, Configurable Parameters - one .xlsx file, matching the real
reference example's tab layout.

Implements the two output-formatting rules confirmed from the reference
images (plan Decision 5): one blank row between consecutive test cases, and
per-test-case metadata / per-phase Test Phase cells merged vertically rather
than repeated on every step row.
"""
from __future__ import annotations

import os
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from .config import Settings
from .schema import TestCase
from .state import PipelineState

_HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
_HEADER_FONT = Font(bold=True)
_TOP_ALIGN = Alignment(vertical="top", wrap_text=True)


def _style_header_row(ws: Worksheet, row: int, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL


def _write_cover_page(wb: Workbook, state: PipelineState, settings: Settings) -> None:
    ws = wb.create_sheet("Cover Page")
    ws.append(["System Qualification Test Specification"])
    ws.append([])
    ws.append(["Project", settings.project_name])
    ws.append(["Version", settings.version])
    ws.append(["Feature", f"{state.get('feature_id', '')} - {state.get('feature_name', '')}"])
    ws.append(["Function Group", state.get("function_group", "")])
    ws.append(["Generated", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")])
    ws.append(["Requirements processed", len(state.get("requirements", []))])
    ws.append(["Test cases generated", len(state.get("test_cases", []))])
    flagged = sum(1 for tc in state.get("test_cases", []) if tc.status == "flagged")
    ws.append(["Test cases flagged for review", flagged])
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 60


def _write_test_pattern(wb: Workbook, state: PipelineState) -> None:
    ws = wb.create_sheet("Test Pattern")
    patterns = state.get("test_patterns", {})

    fixed_names: list[str] = []
    variable_names: list[str] = []
    for rows in patterns.values():
        for row in rows:
            for name in row.fixed_values:
                if name not in fixed_names:
                    fixed_names.append(name)
            for name in row.variable_transitions:
                if name not in variable_names:
                    variable_names.append(name)

    header = ["Requirement ID", "Test Case No."] + fixed_names + variable_names
    ws.append(header)
    _style_header_row(ws, 1, len(header))

    for req_id, rows in patterns.items():
        for row in rows:
            line = [req_id, row.test_case_no]
            line += [row.fixed_values.get(name, "") for name in fixed_names]
            line += [row.variable_transitions.get(name, "") for name in variable_names]
            ws.append(line)

    for col in range(1, len(header) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 18


def _write_item_list(wb: Workbook, state: PipelineState) -> None:
    ws = wb.create_sheet("Item List")
    header = ["Testcase No", "Test type", "Feature Name", "Test Case ID", "Variant", "Execution Required", "Remarks"]
    ws.append(header)
    _style_header_row(ws, 1, len(header))

    for i, tc in enumerate(state.get("test_cases", []), start=1):
        ws.append([i, "Normal_system", tc.feature, tc.test_case_id, tc.variant or "", "Yes", tc.remarks_summary or ""])

    widths = [12, 16, 20, 16, 10, 16, 40]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width


def _write_configurable_parameters(wb: Workbook, state: PipelineState) -> None:
    ws = wb.create_sheet("Configurable Parameters")
    header = ["Parameter Name", "Description", "A1", "A2", "B", "C1", "C2", "C3", "D1", "D2"]
    ws.append(header)
    _style_header_row(ws, 1, len(header))

    seen: set[str] = set()
    for param in state.get("app_param_valid", []):
        name = param.parameter_name or ""
        if not name or name in seen:
            continue
        seen.add(name)
        ws.append([name, param.parameter_description or "", "", "", "", "", "", "", "", ""])

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 60


_TEST_CASE_COLUMNS = [
    "TestCase ID", "Feature", "Variant", "Requirement IDs", "Priority", "Mode of Execution",
    "Test case Description", "Test Phase", "Step", "Test steps", "Parameter Settings", "Units",
    "Expected Value", "Units", "Whether to execute the command", "Remarks",
]


def _write_test_cases(wb: Workbook, state: PipelineState) -> None:
    ws = wb.create_sheet("Test Cases")
    ws.append(_TEST_CASE_COLUMNS)
    _style_header_row(ws, 1, len(_TEST_CASE_COLUMNS))

    row_idx = 2
    for tc in state.get("test_cases", []):
        if not tc.steps:
            continue
        block_start = row_idx

        phase_start = row_idx
        current_phase = tc.steps[0].phase
        for step in tc.steps:
            if step.phase != current_phase:
                ws.merge_cells(start_row=phase_start, start_column=8, end_row=row_idx - 1, end_column=8)
                ws.cell(row=phase_start, column=8, value=current_phase)
                current_phase = step.phase
                phase_start = row_idx

            ws.cell(row=row_idx, column=9, value=step.step_no)
            ws.cell(row=row_idx, column=10, value=step.step_text)
            ws.cell(row=row_idx, column=11, value=step.parameter_settings or "")
            ws.cell(row=row_idx, column=12, value=step.units or "")
            ws.cell(row=row_idx, column=13, value=step.expected_value or "")
            ws.cell(row=row_idx, column=14, value=step.units2 or "")
            ws.cell(row=row_idx, column=15, value=step.whether_execute)
            ws.cell(row=row_idx, column=16, value=step.remarks or "")
            row_idx += 1

        ws.merge_cells(start_row=phase_start, start_column=8, end_row=row_idx - 1, end_column=8)
        ws.cell(row=phase_start, column=8, value=current_phase)

        block_end = row_idx - 1
        if block_end > block_start:
            for col in range(1, 8):
                ws.merge_cells(start_row=block_start, start_column=col, end_row=block_end, end_column=col)
        ws.cell(row=block_start, column=1, value=tc.test_case_id)
        ws.cell(row=block_start, column=2, value=tc.feature)
        ws.cell(row=block_start, column=3, value=tc.variant or "")
        ws.cell(row=block_start, column=4, value=", ".join(tc.requirement_ids))
        ws.cell(row=block_start, column=5, value=tc.priority or "")
        ws.cell(row=block_start, column=6, value=tc.mode_of_execution)
        description = tc.description
        if tc.status == "flagged" and tc.flag_reason:
            description = f"{description}\n[FLAGGED FOR REVIEW: {tc.flag_reason}]"
        ws.cell(row=block_start, column=7, value=description).alignment = _TOP_ALIGN

        row_idx += 1  # one blank row between consecutive test cases (plan Decision 5)

    widths = [16, 16, 10, 20, 10, 16, 40, 14, 6, 40, 18, 8, 18, 8, 12, 30]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width


def write_output_workbook(state: PipelineState, settings: Settings) -> list[str]:
    """Write the single SYS5 output workbook into settings.output_dir and
    return the list of filenames written (relative to output_dir)."""
    os.makedirs(settings.output_dir, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)
    _write_cover_page(wb, state, settings)
    _write_test_pattern(wb, state)
    _write_item_list(wb, state)
    _write_configurable_parameters(wb, state)
    _write_test_cases(wb, state)

    filename = f"{settings.project_name}_SYS5_{state.get('feature_id', 'feature')}_TestCases.xlsx"
    wb.save(os.path.join(settings.output_dir, filename))
    return [filename]
