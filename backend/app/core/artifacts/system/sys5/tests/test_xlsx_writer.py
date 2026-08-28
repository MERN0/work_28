from __future__ import annotations

import os

from openpyxl import load_workbook

from ..config import Settings
from ..schema import AppParameter, TestCase, TestStep
from ..xlsx_writer import write_output_workbook


def _settings(output_dir: str) -> Settings:
    return Settings(
        project_name="tmhc_demo", username="", version="V1.0", domain="automotive", artifact="SYS5",
        model="", input_folder_path="", output_folder_path=output_dir, output_dir=output_dir,
        uploaded_files=[], req_filename="reqs.xlsx", req_sheet_name="002",
    )


def _step(step_no: int, phase: str, text: str) -> TestStep:
    return TestStep(step_no=step_no, phase=phase, keyword="Set", target_ref=None, ref_kind="none", step_text=text)


def _test_case(tc_id: str) -> TestCase:
    return TestCase(
        test_case_id=tc_id,
        feature="Slope_Assist",
        variant="A1",
        requirement_ids=["TMHC_SYSRS_FR002001"],
        priority="P1",
        description="Check slope assist enables.",
        steps=[
            _step(1, "PRECONDITION", "Test_start"),
            _step(2, "PRECONDITION", "Compound Power_On_A1"),
            _step(3, "ACTION", "Set MDL_SEN_Slope_Angle"),
            _step(4, "ACTION", "Verify CAN_Main_Warning_AutoHMode"),
            _step(5, "POSTCONDITION", "End_of_test"),
        ],
        status="clean",
    )


def test_write_output_workbook_layout(tmp_path):
    output_dir = str(tmp_path / "out")
    settings = _settings(output_dir)
    state = {
        "feature_id": "002",
        "feature_name": "Slope Assist",
        "function_group": "Traction",
        "requirements": [],
        "test_patterns": {},
        "app_param_valid": [AppParameter(parameter_name="Slope_Detection_Latency", parameter_description="Latency param")],
        "test_cases": [_test_case("TMHC_SQTC_1"), _test_case("TMHC_SQTC_2")],
    }

    files = write_output_workbook(state, settings)
    assert len(files) == 1
    path = os.path.join(output_dir, files[0])
    assert os.path.exists(path)

    wb = load_workbook(path)
    assert set(wb.sheetnames) == {"Cover Page", "Test Pattern", "Item List", "Configurable Parameters", "Test Cases"}

    ws = wb["Test Cases"]
    # header + 5 steps for case 1 (rows 2-6) + 1 blank row (7) + 5 steps for case 2 (rows 8-12)
    assert ws.cell(row=2, column=1).value == "TMHC_SQTC_1"
    assert ws.cell(row=6, column=1).value is None  # merged away, not repeated on every row
    assert ws.cell(row=7, column=1).value is None  # the blank separator row (plan Decision 5)
    assert ws.cell(row=7, column=10).value is None
    assert ws.cell(row=8, column=1).value == "TMHC_SQTC_2"

    id_merges = [str(r) for r in ws.merged_cells.ranges if str(r).startswith("A2:A6")]
    assert id_merges, "expected the TestCase ID column merged across all of case 1's step rows"

    phase_merges = {str(r) for r in ws.merged_cells.ranges}
    assert "H2:H3" in phase_merges  # PRECONDITION spans rows 2-3
    assert "H4:H5" in phase_merges  # ACTION spans rows 4-5

    item_list = wb["Item List"]
    assert item_list.cell(row=2, column=4).value == "TMHC_SQTC_1"
    assert item_list.cell(row=3, column=4).value == "TMHC_SQTC_2"

    params = wb["Configurable Parameters"]
    assert params.cell(row=2, column=1).value == "Slope_Detection_Latency"
    assert params.cell(row=2, column=3).value in ("", None)  # variant columns left blank
