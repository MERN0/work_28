"""End-to-end test of sys5.generate(): feature_index -> requirements_extract
-> test_pattern_gen -> JSON file, with no real LLM/network call (call_llm is
stubbed at its call site in sys5.py). Self-contained: builds its own tiny
synthetic System Requirements workbook rather than depending on shared test
fixtures/conftest infrastructure.
"""
from __future__ import annotations

import json

import pytest
from openpyxl import Workbook

from .. import sys5

FEATURE_ID = "002"


@pytest.fixture()
def requirements_workbook(tmp_path) -> str:
    wb = Workbook()
    wb.remove(wb.active)

    index = wb.create_sheet("Index")
    index.append(["Feature ID Link", "Feature Name", "Function Group"])
    index.append(["002", "Slope Assist", "Traction"])

    req = wb.create_sheet(FEATURE_ID)
    req.append(
        [
            "Requirement ID", "Requirement Description", "Category", "Variant", "Priority",
            "Verification Method", "Verification Criteria", "Verification Stage", "Source",
            "Status", "Release", "Downstream Traceability", "Remarks",
        ]
    )
    req.append(["", "Slope Assist Requirements", "Heading", "", "", "", "", "", "", "", "", "", ""])
    req.append(
        [
            "TMHC_SYSRS_FR002001",
            "The system shall enable slope assist when the slope angle exceeds the threshold while moving forward.",
            "Functional Requirement", "Variant 1", "P1",
            "Testing", "Verify slope assist enables when slope angle transitions from 0 deg to 3 deg while moving forward in E mode.",
            "System Testing", "CustReq-1", "Approved", "R1", "", "",
        ]
    )

    path = tmp_path / "System Requirements.xlsx"
    wb.save(path)
    return str(path)


def _fake_call_llm(llm, system_prompt, user_input, output_schema, pipeline_config=None):
    return sys5._PatternPlan(
        scenarios=[
            sys5._ScenarioPlan(
                scenario_id="enable-on-slope",
                variable_transitions=[sys5._FactorTransition(factor_name="Option Set", transition="Disabled -> Enabled")],
                applicable_fixed_factor_names=["Direction Switch"],
            )
        ]
    )


def test_generate_writes_json_with_test_patterns_for_every_requirement(requirements_workbook, tmp_path, monkeypatch):
    monkeypatch.setattr(sys5, "call_llm", _fake_call_llm)

    output_dir = tmp_path / "out"
    config = {
        "project_name": "UnitTest",
        "input_folder_path": "",
        "output_dir": str(output_dir),
        "uploaded_files": [requirements_workbook],
        "req_filename": "System Requirements.xlsx",
        "req_sheet_name": FEATURE_ID,
    }

    result_path = sys5.generate(config)

    assert result_path.endswith(".json")
    with open(result_path) as fh:
        payload = json.load(fh)

    assert payload["feature_id"] == FEATURE_ID
    assert payload["feature_name"] == "Slope Assist"
    # Only the one Functional Requirement row - the Heading row is dropped.
    assert len(payload["requirements"]) == 1

    req = payload["requirements"][0]
    assert req["req_id"] == "TMHC_SYSRS_FR002001"
    assert req["category"] == "Functional Requirement"
    # Direction Switch has 2 values (FWD, BWD) -> 2 combinatorial rows.
    assert len(req["test_patterns"]) == 2
    assert {p["fixed_values"]["Direction Switch"] for p in req["test_patterns"]} == {"FWD", "BWD"}


def test_missing_factor_table_fails_fast(requirements_workbook, tmp_path):
    output_dir = tmp_path / "out"
    config = {
        "project_name": "UnitTest",
        "input_folder_path": "",
        "output_dir": str(output_dir),
        "uploaded_files": [requirements_workbook],
        "req_filename": "System Requirements.xlsx",
        "req_sheet_name": "999",  # not registered in _FACTOR_TABLES, and not in the Index sheet either
    }

    with pytest.raises(ValueError, match="not found in the Index sheet"):
        sys5.generate(config)
