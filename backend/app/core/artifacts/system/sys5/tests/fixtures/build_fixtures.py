"""Builds small synthetic xlsx fixtures matching the full documented input
schema (not just what a reference screenshot happened to show - see the
plan's Verification plan section). Used by pytest fixtures in conftest.py.
"""
from __future__ import annotations

import os

from openpyxl import Workbook

FEATURE_ID = "002"


def _save(wb: Workbook, path: str) -> str:
    wb.remove(wb.active)
    wb.save(path)
    return path


def build_system_requirements(out_dir: str) -> str:
    wb = Workbook()

    cover = wb.create_sheet("Cover Page")
    cover.append(["System Requirements - Test Fixture"])

    index = wb.create_sheet("Index")
    index.append(["Feature ID Link", "Feature Name", "Function Group"])
    index.append(["002", "Slope Assist", "Traction"])
    index.append(["005", "Other Feature", "Braking"])

    abbr = wb.create_sheet("Master List Abbreviations")
    abbr.append(["Abbreviations", "Description/Definition"])
    abbr.append(["MB Contactor", "Main Battery Contactor"])
    abbr.append(["RPM", "Rotations per minute"])

    req = wb.create_sheet(FEATURE_ID)
    req.append(
        [
            "Requirement ID", "Requirement Description", "Category", "Variant", "Priority",
            "Verification Method", "Verification Criteria", "Verification Stage", "Source",
            "Status", "Release", "Downstream Traceability", "Remarks",
        ]
    )
    req.append(["", "Slope Assist Requirements", "Heading", "", "", "", "", "", "", "", "", "", ""])
    req.append(
        [
            "TMHC_SYSRS_FR002001",
            "The system shall enable slope assist when the slope angle exceeds the threshold while moving forward.",
            "Functional Requirement", "Variant 1", "P1",
            "Testing", "Verify slope assist enables when slope angle transitions from 0 deg to 3 deg while moving forward in E mode.",
            "System Testing", "CustReq-1", "Approved", "R1", "", "",
        ]
    )
    req.append(
        [
            "TMHC_SYSRS_FR002002",
            "The system shall disable slope assist when the option set is disabled.",
            "Funtional Requiremnt",  # typo'd category - deliberately NOT a clean fuzzy match against the known vocabulary
            "Variant 1", "P2", "Testing", "Verify slope assist disables when Option Set is Disabled.",
            "System Testing", "CustReq-2", "Approved", "R1", "", "",
        ]
    )

    comm = wb.create_sheet("Master Comm Matrix (CAN)")
    comm_headers = [
        "Signal ID", "Message Name", "Message IDs", "Logical Signal Name", "Signal name", "Signal Description",
        "Message DLC", "Periodicity", "Length", "Bit Positions (0 to 63)", "Resolution", "Physical Range",
        "Integer Range", "Unit", "Default Value", "Start up Value", "Message_Send Type",
        "ECU HW (Transmitting)", "ECU HW (Receiving)", "Remarks", "001", "002", "005",
    ]
    comm.append(comm_headers)
    comm.append(
        ["TMHC_SYSRS_DBC0001", "Main_SDO_Tx", "0x581", "PwrCtrlMode", "Main_TxS_0x2020_0x01", "Power control mode",
         "8", "Event", "8", "0-7", "", "", "", "", "", "", "Event", "Main", "HIL", "", "x", "O", "x"]
    )
    comm.append(
        ["TMHC_SYSRS_DBC0002", "Main_SDO_Tx", "0x2040", "OptionSetFlag", "Main_TxS_0x2040_0x05", "Option set flag",
         "8", "Event", "8", "0-7", "", "", "", "", "", "", "Event", "Main", "HIL", "", "x", "x", "O"]
    )
    comm.append(
        ["TMHC_SYSRS_DBC0003", "Disp_PDO1_Rx", "0x221", "SlopeWarning", "Disp_Rx1_Warning", "Slope warning signal",
         "8", "Event", "8", "0-7", "", "", "", "", "", "", "Event", "Disp", "Main", "", "x", "?", "x"]
    )

    param = wb.create_sheet("Master List - App Parameter")
    param_headers = [
        "Parameter ID", "Parameter Name", "Parameter Description(purpose)", "Parameter Type", "Unit",
        "Parameter Valid Value", "Parameter default value", "Parameter min Value", "Parameter max Value",
        "Resolution", "Interdependency with other parameters", "Linked System requirements", "Variant",
        "Change by EOL?", "Change by Service?", "Access Conditions", "Remarks", "001", "002",
    ]
    param.append(param_headers)
    param.append(
        ["TMHC_SYSRS_PARM0001", "Slope_Detection_Latency", "Time within which a slope should be detected",
         "Configuration", "ms", "300", "300", "0", "1000", "1", "", "", "Common", "No", "No", "No", "", "x", "O"]
    )

    io_sig = wb.create_sheet("Master Input Output Signals")
    io_headers = [
        "Signal ID", "Logical Signal Name", "Signal Type", "Variants", "ECU", "Input/Output",
        "Maximum Rated Voltage(V)", "Nominal Operating current(mA)", "Minimum Rated Voltage(V)", "Remarks",
        "001", "002",
    ]
    io_sig.append(io_headers)
    io_sig.append(
        ["TMHC_SYSRS_IO0001", "Slope_Sensor", "Sensor", "A1", "Main", "Input", "5", "20", "0", "", "x", "O"]
    )

    return _save(wb, os.path.join(out_dir, "System Requirements.xlsx"))


def build_command_list(out_dir: str) -> str:
    wb = Workbook()
    cl = wb.create_sheet("Command List")
    cl.append(["Sl.No.", "Type", "Command name", "Message Name", "Signal Description", "Signal Name", "Index (Hex)", "SubIndex (Hex)", "Decimal"])
    cl.append([1, "CAN", "CAN_HIL_PwrCtrlMode", "Main_SDO_Tx", "Power control mode", "Main_TxS_0x2020_0x01", "0x2020", "0x01", "8224"])
    cl.append([2, "CAN", "CAN_Main_OptionSetFlag", "Main_SDO_Tx", "Option set flag", "Main_TxS_0x2040_0x05", "0x2040", "0x05", "8261"])
    return _save(wb, os.path.join(out_dir, "TE_TMHC_Command_List.xlsx"))


def build_configuration_file(out_dir: str) -> str:
    wb = Workbook()
    tol = wb.create_sheet("Tolerances")
    tol.append(["Sl No", "Tolerance Configuration", "Description", "Example", "Unit", "Value (+,-)", "Tolerance Unit", "Remarks"])
    tol.append([1, "Config_Tol_Spd", "Tolerance for Vehicle speed", "4 (Config_Tol_Spd)", "km/hr", "0.5,1.5", "units", "First value +, second -"])
    tol.append([2, "Config_Tol_Deg", "Tolerance for Tilt angle", "0 (Config_Tol_Deg)", "deg", "1,1", "deg", "First value +, second -"])

    mim = wb.create_sheet("Model_Input_Mapping")
    mim.append(["Sl.No.", "Signal", "Test Case Input", "Model Input", "Model Output to ECU", "Remark"])
    # Simulate merged Signal/Sl.No cells: only the first row of each group carries a value.
    mim.append([1, "MDL_SWH_DIR_STATE", "FWD", "1", "0V (DSF) - Low / 5V (DSR)", "Forward"])
    mim.append([None, None, "NEUTRAL", "2", "5V (DSF) - High / 5V (DSR)", "Neutral"])
    mim.append([None, None, "BWD", "3", "5V (DSF) - High / 0V (DSR)", "Backward"])
    mim.append([2, "MDL_SEN_Slope_Angle", "0 deg", "0", "", ""])
    mim.append([None, None, "3 deg", "3", "", ""])
    return _save(wb, os.path.join(out_dir, "TE_TMHC_Configuration_File.xlsx"))


def build_compound_commands(out_dir: str) -> str:
    wb = Workbook()
    set_ws = wb.create_sheet("Compound Commands (Set)")
    set_ws.append(["Compound Power_On_A1"])
    set_ws.append(["Set MDL_PS_B48V", "Parameter Settings", "Units", "Expected Value", "Units", "Whether to execute the command", "Remarks"])
    set_ws.append(["Set MDL_PS_B48V", "52", "V", "", "", "Yes", "Powering up the ECU with battery voltage"])
    set_ws.append(["Wait", "2000", "ms", "", "", "Yes", "Time delay is given as 2000 ms"])
    set_ws.append(["Compound Key_On_A1"])
    set_ws.append(["Set MDL_PS_VBKY", "Parameter Settings", "Units", "Expected Value", "Units", "Whether to execute the command", "Remarks"])
    set_ws.append(["Set MDL_PS_VBKY", "ON", "", "", "", "Yes", "Switch input of key"])

    verify_ws = wb.create_sheet("Compound Commands (Verify)")
    verify_ws.append(["Compound Verifying_Power_On_A1"])
    verify_ws.append(["Verify MDL_PS_B48V", "Parameter Settings", "Units", "Expected Value", "Units", "Whether to execute the command", "Remarks"])
    verify_ws.append(["Verify MDL_PS_B48V", "", "", "52", "V", "Yes", "Electric source through battery"])
    return _save(wb, os.path.join(out_dir, "TE_TMHC_Compound_Commands.xlsx"))


def build_keyword_library(out_dir: str) -> str:
    wb = Workbook()
    lib = wb.create_sheet("Library List")
    lib.append(["Sl.no", "Library", "Library Description", "Example Usage"])
    lib.append([1, "Lib_Ramp Signal_Name(Start=X,Stop=X,Step=X,Time=X)", "Ramp a signal value from Start to Stop.", "Lib_Ramp MDL_SEN_Accelerator_Pedal(Start=0,Stop=80,Step=5,Time=1)"])
    lib.append([2, "Lib_CheckTorqueLimit (Map=MapX,Discharge=X,PCM=X,Delay=X,Read=CAN_ReadSignal,Verify=CAN_VerifySignal)", "Verify torque current within Map List limits.", ""])

    kw = wb.create_sheet("Custom Keyword&Library Details")
    kw.append(["Sl.no", "Format", "Type", "Logical Formula / Method", "Example Reference"])
    kw.append([1, "Lib_Ramp Signal_Name(Start=X,Stop=X,Step=X,Time=X)", "Custom Library", "Ramp Signal_Name from Start to Stop by Step every Time ms.", "Lib_Ramp MDL_SEN_Accelerator_Pedal(Start=0,Stop=80,Step=5,Time=1)"])
    return _save(wb, os.path.join(out_dir, "TE_TMHC_HILLS_Development & Testing_Keyword_Library_Description_Sheet.xlsx"))


def build_all(out_dir: str) -> dict[str, str]:
    os.makedirs(out_dir, exist_ok=True)
    return {
        "requirements": build_system_requirements(out_dir),
        "command_list": build_command_list(out_dir),
        "configuration": build_configuration_file(out_dir),
        "compound_commands": build_compound_commands(out_dir),
        "keyword_library": build_keyword_library(out_dir),
    }
