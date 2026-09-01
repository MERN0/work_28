"""Entry point for the SYS5 artifact: extracts every valid Test Pattern for
every available Functional Requirement of one feature
(config["req_sheet_name"]) from a System Requirements workbook, and saves the
result as a single JSON file.

Everything the pipeline needs lives in this one file (plus `cli.py`, the
command-line wrapper around `generate()`, and `pipeline_config.json`, the
editable engineering-knobs file `PipelineConfig.load()` reads) - no other
module in this package. The file is organized top-to-bottom as the pipeline
actually runs: logging -> config -> Excel I/O -> data schema -> workbook
store -> factor tables -> LLM client -> prompts -> pipeline stages -> the
`generate()`/`main()` entry point.
"""
from __future__ import annotations

import dataclasses
import itertools
import json
import logging
import os
import re
import sys
import time
from contextlib import contextmanager
from dataclasses import dataclass, field, fields
from datetime import datetime
from typing import Any, Iterator, Optional, Type, TypeVar

import openpyxl
from langchain_core.language_models import BaseChatModel
from langchain_core.messages import HumanMessage, SystemMessage
from langchain_openai import ChatOpenAI
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel, Field, ValidationError
from rapidfuzz import fuzz, process

if __package__ in (None, ""):
    # Running as a standalone script (`python sys5.py ...`), not imported as
    # part of the app.core.artifacts.system.sys5 package - there's no package
    # context for the relative import in cli.py (or a future caller) to
    # resolve against. Locate backend/ (this file's 5th ancestor: sys5/ ->
    # system/ -> artifacts/ -> core/ -> app/ -> backend/), put it on
    # sys.path, and set __package__ (PEP 366) so relative imports work
    # exactly as they do when the package is imported normally.
    sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), *([os.pardir] * 5))))
    __package__ = "app.core.artifacts.system.sys5"

T = TypeVar("T", bound=BaseModel)

_ROOT_LOGGER_NAME = "sys5"
_DEFAULT_PIPELINE_CONFIG_PATH = os.path.join(os.path.dirname(__file__), "pipeline_config.json")


# ============================================================================
# Logging
# ============================================================================
# Every stage below logs through get_logger(__name__ analog)/stage_timer():
# "-> stage (context)" on entry, "<- stage done in Xs" on success, or a full
# exception on failure. Output goes to the console and to
# <output_dir>/sys5_run.log.

def configure_logging(pipeline_config: "PipelineConfig", output_dir: str | None = None) -> logging.Logger:
    """Configure the `sys5` logger tree: console output always, plus a
    per-run log file inside `output_dir` (so it ships alongside the
    generated JSON for later debugging) when `pipeline_config.log_to_file` is
    set. Safe to call more than once - handlers are replaced, not stacked."""
    root = logging.getLogger(_ROOT_LOGGER_NAME)
    root.setLevel(pipeline_config.log_level)
    root.propagate = False
    for handler in list(root.handlers):
        root.removeHandler(handler)

    formatter = logging.Formatter(pipeline_config.log_format)

    for stream in (sys.stdout, sys.stderr):
        try:
            stream.reconfigure(errors="backslashreplace")
        except (AttributeError, ValueError):
            pass

    console = logging.StreamHandler()
    console.setFormatter(formatter)
    root.addHandler(console)

    if pipeline_config.log_to_file and output_dir:
        os.makedirs(output_dir, exist_ok=True)
        file_handler = logging.FileHandler(
            os.path.join(output_dir, pipeline_config.log_file_name), encoding="utf-8", errors="backslashreplace"
        )
        file_handler.setFormatter(formatter)
        root.addHandler(file_handler)

    unknown = getattr(pipeline_config, "_unknown_keys", None)
    if unknown:
        root.warning("pipeline_config.json has unrecognized key(s), ignored: %s", sorted(unknown))

    return root


def get_logger(name: str) -> logging.Logger:
    if not (name == _ROOT_LOGGER_NAME or name.startswith(_ROOT_LOGGER_NAME + ".")):
        name = f"{_ROOT_LOGGER_NAME}.{name}"
    return logging.getLogger(name)


@contextmanager
def stage_timer(logger: logging.Logger, stage: str, **context) -> Iterator[None]:
    """Logs `-> stage (ctx...)` on entry and `<- stage done in Xs` on a clean
    exit, or `x  stage FAILED after Xs` (with traceback) if the block raises."""
    ctx = " ".join(f"{k}={v!r}" for k, v in context.items() if v is not None)
    suffix = f" ({ctx})" if ctx else ""
    logger.info("-> %s%s", stage, suffix)
    start = time.monotonic()
    try:
        yield
    except Exception:
        logger.exception("x  %s FAILED after %.1fs%s", stage, time.monotonic() - start, suffix)
        raise
    else:
        logger.info("<- %s done in %.1fs%s", stage, time.monotonic() - start, suffix)


_logger = get_logger(__name__)


# ============================================================================
# Pipeline configuration (engineering knobs - see pipeline_config.json)
# ============================================================================
# Load order (later wins): dataclass defaults -> pipeline_config.json (path
# from `SYS5_PIPELINE_CONFIG_PATH` env var, else the file next to this
# module) -> a small set of env vars for the LLM connection specifically
# (SYS5_LLM_MODEL, SYS5_LLM_API_KEY, SYS5_LLM_API_BASE, SYS5_LLM_MAX_RETRIES,
# SYS5_LLM_TIMEOUT).

@dataclass
class PipelineConfig:
    # -- LLM connection & call resilience -----------------------------------
    llm_model: str = "llm-1-gpt-osx-120b"
    llm_api_key: str = "sk-dfK6wRAt7vIiphRybrrdJQ"
    llm_api_base: str = "http://10.1.2.186:4000"
    llm_temperature: float = 0
    llm_max_retries: int = 3
    llm_timeout_seconds: int = 300
    structured_output_max_retries: int = 2
    structured_output_method: str = "json_schema"
    llm_reasoning_effort: str | None = None
    llm_output_version: str = "v0"

    # -- Fuzzy-matching thresholds (0-100, higher = stricter) ---------------
    header_row_match_threshold: int = 75
    column_match_threshold: int = 75
    sheet_name_match_threshold: int = 80
    category_match_threshold: int = 95

    # -- Performance ----------------------------------------------------------
    max_test_cases_per_requirement: int = 5

    # -- Logging -----------------------------------------------------------
    log_level: str = "INFO"
    log_to_file: bool = True
    log_file_name: str = "sys5_run.log"
    log_format: str = "%(asctime)s %(levelname)-7s %(name)s: %(message)s"

    @classmethod
    def load(cls, path: str | None = None) -> "PipelineConfig":
        path = path or os.environ.get("SYS5_PIPELINE_CONFIG_PATH", _DEFAULT_PIPELINE_CONFIG_PATH)
        data: dict = {}
        if path and os.path.isfile(path):
            with open(path) as fh:
                data = json.load(fh)

        valid_fields = {f.name for f in fields(cls)}
        unknown = set(data) - valid_fields
        kwargs = {k: v for k, v in data.items() if k in valid_fields}
        config = cls(**kwargs)
        if unknown:
            config._unknown_keys = unknown  # type: ignore[attr-defined]

        env_overrides = {
            "llm_model": os.environ.get("SYS5_LLM_MODEL"),
            "llm_api_key": os.environ.get("SYS5_LLM_API_KEY"),
            "llm_api_base": os.environ.get("SYS5_LLM_API_BASE"),
        }
        for key, value in env_overrides.items():
            if value:
                setattr(config, key, value)
        if os.environ.get("SYS5_LLM_MAX_RETRIES"):
            config.llm_max_retries = int(os.environ["SYS5_LLM_MAX_RETRIES"])
        if os.environ.get("SYS5_LLM_TIMEOUT"):
            config.llm_timeout_seconds = int(os.environ["SYS5_LLM_TIMEOUT"])

        return config

    def as_dict(self) -> dict:
        return dataclasses.asdict(self)


# ============================================================================
# Settings (the per-run `config` dict handed to generate())
# ============================================================================

@dataclass
class Settings:
    project_name: str
    username: str
    version: str
    domain: str
    artifact: str
    model: str
    input_folder_path: str
    output_folder_path: str
    output_dir: str
    uploaded_files: list
    req_filename: str
    req_sheet_name: str

    @classmethod
    def from_config(cls, config: dict) -> "Settings":
        return cls(
            project_name=config["project_name"],
            username=config.get("username", ""),
            version=config.get("version", ""),
            domain=config.get("domain", ""),
            artifact=config.get("artifact", "SYS5"),
            model=config.get("model", ""),
            input_folder_path=config["input_folder_path"],
            output_folder_path=config.get("output_folder_path", config.get("output_dir", "")),
            output_dir=config["output_dir"],
            uploaded_files=config.get("uploaded_files", []) or [],
            req_filename=config["req_filename"],
            req_sheet_name=config["req_sheet_name"],
        )


# ============================================================================
# Excel I/O helpers (judgment-free: read cells honestly, fuzzy-locate
# sheets/headers/columns under typo/whitespace variance)
# ============================================================================

def load_workbook(path: str):
    return openpyxl.load_workbook(path, data_only=True, read_only=True)


def _norm(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def find_sheet(wb, *candidates: str, threshold: int = 80) -> Optional[Worksheet]:
    """Fuzzy-match a sheet name against `candidates`, tried in order."""
    names = wb.sheetnames
    norm_names = [_norm(n).lower() for n in names]
    for cand in candidates:
        norm_cand = _norm(cand).lower()
        if norm_cand in norm_names:
            return wb[names[norm_names.index(norm_cand)]]
        match = process.extractOne(norm_cand, norm_names, scorer=fuzz.token_sort_ratio)
        if match and match[1] >= threshold:
            return wb[names[norm_names.index(match[0])]]
    return None


def sheet_matrix(ws: Worksheet, max_rows: Optional[int] = None) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        rows.append(list(row))
        if max_rows and i + 1 >= max_rows:
            break
    return rows


def find_header_row(
    matrix: list[list[Any]],
    expected_anchors: list[str],
    max_scan_rows: int = 15,
    threshold: int = 75,
) -> Optional[int]:
    """Best-guess 0-based index of the header row, by fuzzy-matching cell text
    in each of the first `max_scan_rows` rows against `expected_anchors`.
    Handles title rows / merged banner rows sitting above the real header."""
    best_idx, best_score = None, 0
    for i, row in enumerate(matrix[:max_scan_rows]):
        cells = [_norm(c) for c in row if _norm(c)]
        if not cells:
            continue
        lowered = [c.lower() for c in cells]
        score = 0
        for anchor in expected_anchors:
            match = process.extractOne(anchor.lower(), lowered, scorer=fuzz.token_sort_ratio)
            if match and match[1] >= threshold:
                score += 1
        if score > best_score:
            best_score, best_idx = score, i
    return best_idx


def resolve_columns(headers: list[Any], candidate_names: list[str], threshold: int = 75) -> dict[str, Optional[int]]:
    """Map each of `candidate_names` to the best-matching column index in
    `headers`, or None if nothing crosses `threshold`."""
    norm_headers = [_norm(h).lower() for h in headers]
    result: dict[str, Optional[int]] = {}
    for cand in candidate_names:
        cand_l = cand.lower()
        if cand_l in norm_headers:
            result[cand] = norm_headers.index(cand_l)
            continue
        match = process.extractOne(cand_l, norm_headers, scorer=fuzz.token_sort_ratio)
        result[cand] = norm_headers.index(match[0]) if match and match[1] >= threshold else None
    return result


_FEATURE_ID_RE = re.compile(r"(\d{1,3})")


def normalize_feature_id(value: Any) -> Optional[str]:
    """Normalize a feature-id-ish value ('005', '5', ' Feature_005 ', 5) to a
    zero-padded 3-digit string, or None if it doesn't contain digits."""
    s = _norm(value)
    if not s:
        return None
    match = _FEATURE_ID_RE.search(s)
    if not match:
        return None
    return match.group(1).zfill(3)


def rows_as_dicts(
    matrix: list[list[Any]], header_row_idx: int, col_map: dict[str, Optional[int]]
) -> list[dict[str, Any]]:
    """Convert every non-blank data row below `header_row_idx` into a dict
    keyed by `col_map`'s keys, reading only real cell values."""
    out: list[dict[str, Any]] = []
    for row in matrix[header_row_idx + 1 :]:
        if all(_norm(c) == "" for c in row):
            continue
        record = {key: (row[idx] if idx is not None and idx < len(row) else None) for key, idx in col_map.items()}
        out.append(record)
    return out


def _fuzzy_key(value: Any) -> str:
    """Normalized form used only for fuzzy scoring: underscores treated as
    word separators too."""
    return _norm(value).lower().replace("_", " ")


def fuzzy_find(needle: Any, haystack: list[str], threshold: int = 90) -> Optional[str]:
    """Return the entry in `haystack` that best fuzzy-matches `needle`, or
    None if nothing crosses `threshold`. Used by requirements_extract's
    Category classification - a value that doesn't cross the threshold
    against the known vocabulary is dropped rather than guessed."""
    n = _fuzzy_key(needle)
    if not n:
        return None
    keyed = [_fuzzy_key(h) for h in haystack]
    if n in keyed:
        return haystack[keyed.index(n)]
    match = process.extractOne(n, keyed, scorer=fuzz.token_sort_ratio)
    return haystack[keyed.index(match[0])] if match and match[1] >= threshold else None


# ============================================================================
# Data schema (pydantic models)
# ============================================================================

class Requirement(BaseModel):
    req_id: str
    description: str
    category: str
    variant: Optional[str] = None
    priority: Optional[str] = None
    verification_method: Optional[str] = None
    verification_criteria: Optional[str] = None
    verification_stage: Optional[str] = None
    source: Optional[str] = None
    status: Optional[str] = None
    release: Optional[str] = None
    downstream_traceability: Optional[str] = None
    remarks: Optional[str] = None


class HeadingInfoRow(BaseModel):
    """A Heading/Information-category row from a requirement sheet, kept as
    queryable context (not turned into a test pattern)."""
    req_id: Optional[str] = None
    description: str
    category: str


def format_heading_info(rows: list["HeadingInfoRow"]) -> str:
    """Render Heading/Information rows as plain text for inclusion in the
    test_pattern_gen prompt. Returns "" for an empty list so callers can
    splice it into an f-string unconditionally."""
    if not rows:
        return ""
    return "\n".join(f"- [{r.category}]{f' ({r.req_id})' if r.req_id else ''} {r.description}" for r in rows)


class Factor(BaseModel):
    name: str
    values: list[str]
    ease_of_adjustment: Optional[str] = None
    signal_ref: Optional[str] = None  # explicit signal this factor sets, if known


class FactorTable(BaseModel):
    feature_id: str
    fixed_factors: list[Factor]
    variable_factors: list[Factor]


class TestPatternRow(BaseModel):
    test_case_no: int
    scenario_id: str
    fixed_values: dict[str, str]     # factor name -> value, for this row
    variable_transitions: dict[str, str]  # factor name -> "A -> B" transition text


# ============================================================================
# Workbook store: parses the System Requirements workbook once into memory
# ============================================================================

class MissingInputFileError(RuntimeError):
    pass


def resolve_requirements_file(input_folder_path: str, req_filename: str, uploaded_files: list[str]) -> str:
    """Locate the System Requirements workbook among `uploaded_files` and/or
    `input_folder_path`: an exact `req_filename` basename match, or (if
    there's exactly one .xlsx/.xlsm candidate) that one file."""
    candidates: list[str] = list(uploaded_files or [])
    if input_folder_path and os.path.isdir(input_folder_path):
        for fname in os.listdir(input_folder_path):
            if fname.lower().endswith((".xlsx", ".xlsm")):
                candidates.append(os.path.join(input_folder_path, fname))
    candidates = sorted(set(candidates))
    if not candidates:
        raise MissingInputFileError(f"No requirements workbook found under {input_folder_path!r} or uploaded_files")

    if req_filename:
        for path in candidates:
            if os.path.basename(path) == req_filename:
                return path

    if len(candidates) == 1:
        return candidates[0]

    raise MissingInputFileError(
        f"Could not identify the requirements workbook among {candidates} - pass the requirements file "
        f"explicitly, or ensure req_filename ({req_filename!r}) matches exactly one candidate's basename."
    )


@dataclass
class InMemoryWorkbookStore:
    file_path: str
    pipeline_config: PipelineConfig = field(default_factory=PipelineConfig)

    feature_index: dict[str, dict[str, str]] = field(default_factory=dict)

    requirement_headers: list[Any] = field(default_factory=list)
    requirement_matrix: list[list[Any]] = field(default_factory=list)
    requirement_header_row: int = -1

    @classmethod
    def load(
        cls, file_path: str, req_sheet_name: str, pipeline_config: Optional[PipelineConfig] = None
    ) -> "InMemoryWorkbookStore":
        store = cls(file_path=file_path, pipeline_config=pipeline_config or PipelineConfig())
        with stage_timer(_logger, "load_inputs: requirements workbook"):
            store._load_requirements_workbook(req_sheet_name)
        _logger.info(
            "workbook store loaded: %d requirement row(s) for sheet %r",
            max(len(store.requirement_matrix) - 1, 0), req_sheet_name,
        )
        return store

    def _find_sheet(self, wb, *candidates: str):
        return find_sheet(wb, *candidates, threshold=self.pipeline_config.sheet_name_match_threshold)

    def _find_header_row(self, matrix: list[list[Any]], anchors: list[str]):
        return find_header_row(matrix, anchors, threshold=self.pipeline_config.header_row_match_threshold)

    def _resolve_columns(self, headers: list[Any], candidates: list[str]) -> dict[str, Optional[int]]:
        return resolve_columns(headers, candidates, threshold=self.pipeline_config.column_match_threshold)

    def _load_sheet(
        self, wb, candidates: list[str], anchors: list[str]
    ) -> Optional[tuple[list[list[Any]], int, list[Any]]]:
        """Find a sheet by name (fuzzy) and its header row (fuzzy, falling
        back to row 0 if nothing scores well against `anchors`)."""
        ws = self._find_sheet(wb, *candidates)
        if ws is None:
            return None
        matrix = sheet_matrix(ws)
        header_row = self._find_header_row(matrix, anchors) or 0
        headers = matrix[header_row] if header_row < len(matrix) else []
        return matrix, header_row, headers

    def _load_requirements_workbook(self, req_sheet_name: str) -> None:
        wb = load_workbook(self.file_path)

        loaded = self._load_sheet(wb, ["Index"], ["Feature ID Link", "Feature Name", "Function Group"])
        if loaded:
            matrix, header_row, headers = loaded
            col_map = self._resolve_columns(headers, ["Feature ID Link", "Feature Name", "Function Group"])
            for row in rows_as_dicts(matrix, header_row, col_map):
                fid = normalize_feature_id(row.get("Feature ID Link"))
                if fid:
                    self.feature_index[fid] = {
                        "name": _norm(row.get("Feature Name")),
                        "function_group": _norm(row.get("Function Group")),
                    }

        loaded = self._load_sheet(wb, [req_sheet_name], ["Requirement ID", "Requirement Description", "Category"])
        if loaded:
            self.requirement_matrix, self.requirement_header_row, self.requirement_headers = loaded

    def get_feature_info(self, feature_id: str) -> Optional[dict[str, str]]:
        return self.feature_index.get(normalize_feature_id(feature_id) or feature_id)

    def get_requirement_rows(self) -> list[dict[str, Any]]:
        col_map = self._resolve_columns(
            self.requirement_headers,
            [
                "Requirement ID", "Requirement Description", "Category", "Variant", "Priority",
                "Verification Method", "Verification Criteria", "Verification Stage", "Source",
                "Status", "Release", "Downstream Traceability", "Remarks",
            ],
        )
        return rows_as_dicts(self.requirement_matrix, self.requirement_header_row, col_map)


# ============================================================================
# Factor tables (human-supplied per-feature domain knowledge)
# ============================================================================

class MissingFactorTableError(RuntimeError):
    """Raised when generate() is asked to process a feature with no
    registered factor table - fails fast rather than letting an LLM invent
    domain knowledge that is explicitly human/expert-owned."""


def _slope_assist_fixed_factors() -> list[Factor]:
    return [
        Factor(name="Truck Size", values=["1t", "3t"], ease_of_adjustment="Reset Required"),
        Factor(name="Discharge Capacity", values=["25% Discharge"], ease_of_adjustment="Easy"),
        Factor(
            name="Power Control Mode",
            values=["P", "S", "E"],
            ease_of_adjustment="Easy",
            signal_ref="CAN_HIL_PwrCtrlMode",
        ),
        Factor(
            name="Direction Switch",
            values=["FWD", "BWD"],
            ease_of_adjustment="Easy",
            signal_ref="MDL_SWH_DIR_STATE",
        ),
        Factor(name="Load Capacity", values=["NL"], ease_of_adjustment="Easy", signal_ref="MDL_SEN_Load"),
    ]


def _slope_assist_variable_factors() -> list[Factor]:
    return [
        Factor(name="Option Set", values=["Disabled", "Enabled"], ease_of_adjustment="Moderately Difficult"),
        Factor(
            name="Slope angle",
            values=["0 deg", "3 deg"],
            ease_of_adjustment="Easy",
            signal_ref="MDL_SEN_Slope_Angle",
        ),
    ]


def _slope_assist_table(feature_id: str) -> FactorTable:
    return FactorTable(
        feature_id=feature_id,
        fixed_factors=_slope_assist_fixed_factors(),
        variable_factors=_slope_assist_variable_factors(),
    )


# Slope Assist's factor table, registered under both feature ids seen for it:
# "002" (the test spec document's own numbering) and "019" (the actual
# System-Requirements Index-sheet feature id).
_FACTOR_TABLES: dict[str, FactorTable] = {
    "002": _slope_assist_table("002"),
    "019": _slope_assist_table("019"),
}


def get_factor_table(feature_id: str) -> FactorTable:
    table = _FACTOR_TABLES.get(feature_id)
    if table is None:
        raise MissingFactorTableError(
            f"No factor table registered for feature {feature_id!r}. Factor tables are human-supplied "
            f"domain knowledge - add one before generating test patterns for this feature."
        )
    return table


def register_factor_table(table: FactorTable) -> None:
    """Register (or overwrite) a feature's factor table at runtime."""
    _FACTOR_TABLES[table.feature_id] = table


# ============================================================================
# LLM client + single-shot structured-output call
# ============================================================================

def get_llm(pipeline_config: PipelineConfig, model: str | None = None) -> ChatOpenAI:
    """Construct a `ChatOpenAI` client configured from `pipeline_config`.
    `output_version="v0"` / `use_responses_api=False` keep message formatting
    compatible with this deployment's litellm proxy in front of a
    self-hosted, non-OpenAI model (gpt-oss-120b via vLLM) - see
    pipeline_config.llm_output_version's comment."""
    optional: dict = {}
    if pipeline_config.llm_reasoning_effort:
        optional["reasoning_effort"] = pipeline_config.llm_reasoning_effort

    return ChatOpenAI(
        model=model or pipeline_config.llm_model,
        api_key=pipeline_config.llm_api_key,
        base_url=pipeline_config.llm_api_base,
        temperature=pipeline_config.llm_temperature,
        max_retries=pipeline_config.llm_max_retries,
        timeout=pipeline_config.llm_timeout_seconds,
        output_version=pipeline_config.llm_output_version,
        use_responses_api=False,
        **optional,
    )


def call_llm(
    llm: BaseChatModel,
    system_prompt: str,
    user_input: str,
    output_schema: Type[T],
    pipeline_config: Optional[PipelineConfig] = None,
) -> T:
    """Make one single-shot structured-output LLM call and return a typed
    `output_schema` instance - a single HTTP round trip, one SystemMessage +
    one HumanMessage in, one parsed answer out, never a second turn. Retries
    up to `pipeline_config.structured_output_max_retries` times on a pydantic
    validation failure, feeding the error back into the prompt so the model
    can correct itself, before giving up with a RuntimeError."""
    max_retries = pipeline_config.structured_output_max_retries if pipeline_config else 2
    method = pipeline_config.structured_output_method if pipeline_config else "json_schema"
    structured_llm = llm.with_structured_output(output_schema, method=method)

    messages = [SystemMessage(system_prompt), HumanMessage(user_input)]
    last_error: Exception | None = None
    started = time.monotonic()

    for attempt in range(max_retries + 1):
        try:
            result = structured_llm.invoke(messages)
            parsed = result if isinstance(result, output_schema) else output_schema.model_validate(result)
            _logger.info(
                "call_llm(%s) succeeded on attempt %d/%d in %.1fs",
                output_schema.__name__, attempt + 1, max_retries + 1, time.monotonic() - started,
            )
            return parsed
        except (ValidationError, ValueError) as exc:
            last_error = exc
            _logger.warning(
                "call_llm(%s) attempt %d/%d produced an invalid answer: %s",
                output_schema.__name__, attempt + 1, max_retries + 1, exc,
            )
            messages.append(
                HumanMessage(f"Your previous answer was invalid: {exc}\nPlease correct it and answer again.")
            )

    raise RuntimeError(
        f"Failed to obtain a valid {output_schema.__name__} after {max_retries + 1} attempts"
    ) from last_error


# ============================================================================
# Prompts
# ============================================================================

_COMMON_RULES = """
You are working on TMHC (Toyota Material Handling) system qualification test
generation. Follow these rules without exception:

1. Never invent, guess, or complete a signal name, command name, compound
   command name, library call, tolerance name, parameter name, or value.
   Every one of those must come verbatim from the context given to you below,
   which was pulled directly from the actual source workbooks. If what you
   need isn't in that context, say so explicitly instead of making a
   plausible-looking substitute.
2. Handle typos, extra whitespace, and inconsistent casing gracefully when
   matching text - the underlying source data has these. But do not use that
   as license to invent new content; only match against what the context
   below actually contains.
3. Most signal/parameter/requirement text in this domain is heavily
   abbreviated - read it with that in mind, but never let an abbreviation
   you're unsure of become an excuse to invent a name that isn't in the
   context given below.
4. Base every decision strictly on the requirement text and the context given
   below. Do not use outside automotive/HIL-testing knowledge to fill gaps -
   if the context below doesn't say it, it isn't true here.
"""

PROMPTS: dict[str, str] = {
    "test_pattern_gen": _COMMON_RULES + """
You are generating the Test Pattern for one Functional Requirement.

You will be given: the requirement's full text (including its Verification
Criteria field), and the feature's factor table (fixed factors that combine
combinatorially, and variable factors that represent the actual transition
being tested).

Your job, in two steps:

1. Read the Verification Criteria field and identify every DISTINCT testable
   scenario it describes. A scenario is a specific qualitative situation to
   verify (e.g. "slope assist enables when the angle exceeds the threshold
   while moving forward"). If Verification Criteria gives a numeric range,
   treat the boundary/equivalence-class values of that range as defining a
   scenario (or scenarios), per standard equivalence-class testing practice -
   do not enumerate every value in the range.

2. For each scenario, decide which of the feature's variable factors it
   exercises and what transition each undergoes (e.g. "Disabled -> Enabled"),
   then take the FULL combinatorial sweep of the feature's fixed factors that
   are applicable to this requirement's Variant. Each combination becomes one
   Test Pattern row for that scenario. Concatenate all scenarios' rows, in
   order, to form the requirement's complete Test Pattern.

If a fixed or variable factor doesn't apply to this specific requirement
(e.g. because the requirement is scoped to a particular variant or mode),
leave it out of the combinatorics rather than including an irrelevant
dimension.
""",
}


def get_prompt(stage: str) -> str:
    return PROMPTS[stage]


# ============================================================================
# Pipeline stages
#
# `PipelineState` is a plain dict threaded through the three stages below, in
# order: feature_index -> requirements_extract -> test_pattern_gen. No
# framework needed for a fixed 3-step linear sequence with no branching -
# each stage is just a function returning an updated copy of the dict.
# ============================================================================

PipelineState = dict[str, Any]

_KNOWN_CATEGORIES = [
    "Heading",
    "Information",
    "Configuration Requirement",
    "Functional Requirement",
    "NonFunctional Requirement",
    "Security Requirement",
]


def stage_feature_index(store: InMemoryWorkbookStore, state: PipelineState) -> PipelineState:
    """Deterministic: exact Index-sheet lookup for the feature id."""
    feature_id = state["feature_id"]
    info = store.get_feature_info(feature_id)
    if info is None:
        raise ValueError(f"Feature id {feature_id!r} was not found in the Index sheet")
    _logger.info("feature_index: feature=%s name=%r function_group=%r", feature_id, info["name"], info["function_group"])
    return {**state, "feature_name": info["name"], "function_group": info["function_group"]}


def _to_requirement(row: dict) -> Requirement:
    return Requirement(
        req_id=_norm(row.get("Requirement ID")),
        description=_norm(row.get("Requirement Description")),
        category="Functional Requirement",
        variant=_norm(row.get("Variant")) or None,
        priority=_norm(row.get("Priority")) or None,
        verification_method=_norm(row.get("Verification Method")) or None,
        verification_criteria=_norm(row.get("Verification Criteria")) or None,
        verification_stage=_norm(row.get("Verification Stage")) or None,
        source=_norm(row.get("Source")) or None,
        status=_norm(row.get("Status")) or None,
        release=_norm(row.get("Release")) or None,
        downstream_traceability=_norm(row.get("Downstream Traceability")) or None,
        remarks=_norm(row.get("Remarks")) or None,
    )


def stage_requirements_extract(
    store: InMemoryWorkbookStore, pipeline_config: PipelineConfig, state: PipelineState
) -> PipelineState:
    """Deterministic: classify each requirement-sheet row's Category via a
    fuzzy match against the known vocabulary. Only "Functional Requirement"
    rows become testable Requirement objects; "Heading"/"Information" rows
    are kept as queryable background context. A row whose Category doesn't
    cross `category_match_threshold` is dropped rather than guessed."""
    threshold = pipeline_config.category_match_threshold
    with stage_timer(_logger, "requirements_extract"):
        rows = store.get_requirement_rows()

        classified: list[tuple[dict, str]] = []
        dropped = 0
        for row in rows:
            raw_category = row.get("Category")
            match = fuzzy_find(raw_category, _KNOWN_CATEGORIES, threshold=threshold)
            if match:
                classified.append((row, match))
            elif _norm(raw_category):
                dropped += 1

        if dropped:
            _logger.warning(
                "requirements_extract: %d row(s) had a Category value that didn't match any known category "
                "(threshold=%d) and were dropped rather than guessed", dropped, threshold,
            )
        _logger.info("requirement sheet: %d row(s), %d classified, %d dropped", len(rows), len(classified), dropped)

        requirements: list[Requirement] = []
        heading_info: list[HeadingInfoRow] = []
        for row, category in classified:
            if category == "Functional Requirement":
                requirements.append(_to_requirement(row))
            elif category in ("Heading", "Information"):
                heading_info.append(
                    HeadingInfoRow(
                        req_id=_norm(row.get("Requirement ID")) or None,
                        description=_norm(row.get("Requirement Description")),
                        category=category,
                    )
                )
        _logger.info("requirements_extract: %d functional requirement(s), %d heading/info row(s)", len(requirements), len(heading_info))

    return {**state, "requirements": requirements, "heading_info": heading_info}


class _FactorTransition(BaseModel):
    factor_name: str = Field(description="Exact name of one of the feature's variable factors.")
    transition: str = Field(description="The transition under test for that factor, e.g. '0 deg -> 3 deg'.")


class _ExcludedValues(BaseModel):
    factor_name: str = Field(description="Exact name of one of the feature's fixed factors.")
    excluded_values: list[str] = Field(description="Values of that factor to leave OUT of the combinatorial sweep.")


class _ScenarioPlan(BaseModel):
    scenario_id: str = Field(description="Short identifier for this scenario, e.g. 'slope_0_to_3_enabled'.")
    variable_transitions: list[_FactorTransition] = Field(
        description="One entry per variable factor this scenario exercises."
    )
    applicable_fixed_factor_names: list[str] = Field(
        description="Exact names of the fixed factors that apply to this scenario (these get swept combinatorially)."
    )
    excluded_fixed_factor_values: list[_ExcludedValues] = Field(
        default_factory=list,
        description="Optional. Only for fixed-factor values this requirement explicitly rules out.",
    )


class _PatternPlan(BaseModel):
    scenarios: list[_ScenarioPlan]


def _expand(plan: _PatternPlan, table: FactorTable) -> list[TestPatternRow]:
    """The LLM only plans scenarios; this does the actual combinatorial
    expansion (mechanical, not a judgment call)."""
    rows: list[TestPatternRow] = []
    counter = 1
    fixed_by_name = {f.name: f for f in table.fixed_factors}
    for scenario in plan.scenarios:
        excluded = {e.factor_name: e.excluded_values for e in scenario.excluded_fixed_factor_values}
        transitions = {t.factor_name: t.transition for t in scenario.variable_transitions}
        chosen = [fixed_by_name[name] for name in scenario.applicable_fixed_factor_names if name in fixed_by_name]
        value_lists = [
            [v for v in factor.values if v not in excluded.get(factor.name, [])]
            for factor in chosen
        ]
        combos = list(itertools.product(*value_lists)) if value_lists else [()]
        for combo in combos:
            rows.append(
                TestPatternRow(
                    test_case_no=counter,
                    scenario_id=scenario.scenario_id,
                    fixed_values=dict(zip([f.name for f in chosen], combo)),
                    variable_transitions=dict(transitions),
                )
            )
            counter += 1
    return rows


def _cap_rows(rows: list[TestPatternRow], max_rows: int) -> list[TestPatternRow]:
    """Cap the total rows for one requirement to at most `max_rows`,
    round-robining across scenarios first (so a cap smaller than the
    scenario count still keeps at least one row per scenario instead of only
    ever keeping the first scenario's rows), then renumbers test_case_no
    sequentially so kept rows stay a gapless 1..N."""
    if max_rows <= 0 or len(rows) <= max_rows:
        return rows
    by_scenario: dict[str, list[TestPatternRow]] = {}
    for row in rows:
        by_scenario.setdefault(row.scenario_id, []).append(row)
    capped: list[TestPatternRow] = []
    while len(capped) < max_rows and any(by_scenario.values()):
        for bucket in by_scenario.values():
            if len(capped) >= max_rows:
                break
            if bucket:
                capped.append(bucket.pop(0))
    return [row.model_copy(update={"test_case_no": i}) for i, row in enumerate(capped, start=1)]


def stage_test_pattern_gen(
    store: InMemoryWorkbookStore, llm: BaseChatModel, pipeline_config: PipelineConfig, state: PipelineState
) -> PipelineState:
    """LLM plans scenarios (one call per requirement); Python does the actual
    combinatorics (deterministic) - see _expand/_cap_rows."""
    max_rows = pipeline_config.max_test_cases_per_requirement
    table = get_factor_table(state["feature_id"])
    patterns: dict[str, list[TestPatternRow]] = {}

    fixed_desc = "\n".join(f"- {f.name}: {f.values}" for f in table.fixed_factors)
    variable_desc = "\n".join(f"- {f.name}: {f.values}" for f in table.variable_factors)
    heading_context = format_heading_info(state.get("heading_info", []))

    for req in state["requirements"]:
        with stage_timer(_logger, "test_pattern_gen", req=req.req_id):
            prompt = get_prompt("test_pattern_gen")
            user_input = (
                f"Requirement {req.req_id}: {req.description}\n"
                f"Verification Criteria: {req.verification_criteria}\n"
                f"Variant: {req.variant}\n\n"
                f"Feature's fixed factors (combine combinatorially):\n{fixed_desc}\n\n"
                f"Feature's variable factors (the transitions actually under test):\n{variable_desc}"
                + (f"\n\nBackground context from Heading/Information rows on the requirement "
                   f"sheet (not requirements themselves, but may clarify intent):\n{heading_context}"
                   if heading_context else "")
            )
            result = call_llm(llm, prompt, user_input, _PatternPlan, pipeline_config=pipeline_config)
            expanded = _expand(result, table)
            patterns[req.req_id] = _cap_rows(expanded, max_rows)
            if len(patterns[req.req_id]) < len(expanded):
                _logger.info(
                    "test_pattern_gen: req=%s -> %d combinatorial row(s), capped to %d",
                    req.req_id, len(expanded), len(patterns[req.req_id]),
                )
            _logger.info("test_pattern_gen: req=%s -> %d test-pattern row(s)", req.req_id, len(patterns[req.req_id]))

    return {**state, "test_patterns": patterns}


def run_pipeline(settings: Settings, pipeline_config: PipelineConfig | None = None) -> PipelineState:
    pipeline_config = pipeline_config or PipelineConfig.load()
    started = time.monotonic()
    _logger.info("run_pipeline starting: feature=%s model=%s", settings.req_sheet_name, pipeline_config.llm_model)

    with stage_timer(_logger, "load_inputs"):
        file_path = resolve_requirements_file(settings.input_folder_path, settings.req_filename, settings.uploaded_files)
        _logger.info("requirements workbook resolved: %s", file_path)
        store = InMemoryWorkbookStore.load(file_path, settings.req_sheet_name, pipeline_config)

    llm = get_llm(pipeline_config, model=settings.model or None)

    state: PipelineState = {"feature_id": settings.req_sheet_name}
    with stage_timer(_logger, "pipeline"):
        state = stage_feature_index(store, state)
        state = stage_requirements_extract(store, pipeline_config, state)
        state = stage_test_pattern_gen(store, llm, pipeline_config, state)

    _logger.info("run_pipeline finished in %.1fs total", time.monotonic() - started)
    return state


# ============================================================================
# Entry point
# ============================================================================

def _build_payload(state: PipelineState, generated_at: str) -> dict:
    test_patterns = state.get("test_patterns", {})
    return {
        "feature_id": state.get("feature_id", ""),
        "feature_name": state.get("feature_name", ""),
        "function_group": state.get("function_group", ""),
        "generated_at": generated_at,
        "requirements": [
            {
                **req.model_dump(),
                "test_patterns": [row.model_dump() for row in test_patterns.get(req.req_id, [])],
            }
            for req in state.get("requirements", [])
        ],
    }


def generate(config: dict) -> str:
    settings = Settings.from_config(config)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    os.makedirs(settings.output_dir, exist_ok=True)  # required before the log file / JSON output are written

    pipeline_config = PipelineConfig.load()
    configure_logging(pipeline_config, output_dir=settings.output_dir)
    logger = get_logger(__name__)
    logger.info(
        "SYS5 generate() starting: project=%s feature=%s output_dir=%s",
        settings.project_name, settings.req_sheet_name, settings.output_dir,
    )
    started = time.monotonic()

    final_state = run_pipeline(settings, pipeline_config)
    payload = _build_payload(final_state, datetime.now().isoformat(timespec="seconds"))

    json_path = os.path.join(settings.output_dir, f"SYS5_TestPatterns_{settings.project_name}_{timestamp}.json")
    with open(json_path, "w", encoding="utf-8") as fh:
        json.dump(payload, fh, indent=2)

    pattern_count = sum(len(p["test_patterns"]) for p in payload["requirements"])
    summary = (
        f"SYS5 test-pattern extraction complete for feature {payload['feature_id']} ({payload['feature_name']}): "
        f"{len(payload['requirements'])} requirement(s) -> {pattern_count} test pattern row(s) "
        f"in {time.monotonic() - started:.1f}s. Output: {json_path}"
    )
    logger.info(summary)
    print(summary)
    return json_path


def main() -> int:
    """Minimal standalone entry point: `python sys5.py <config.json>`. Reads
    the config dict from the given JSON file, runs generate(), and prints the
    resulting artifact path."""
    if len(sys.argv) < 2:
        print("Usage: python sys5.py <config.json>", file=sys.stderr)
        return 1

    with open(sys.argv[1]) as fh:
        config = json.load(fh)

    print(generate(config))
    return 0


if __name__ == "__main__":
    sys.exit(main())
