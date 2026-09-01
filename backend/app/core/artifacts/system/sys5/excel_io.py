"""Low-level, judgment-free Excel I/O helpers.

Everything here is mechanical: reading real cells honestly, and locating
header rows/columns/sheets under typo/whitespace variance via fuzzy string
matching. None of this decides whether a row is *semantically* valid - that's
workbook_store.py / the node layer. This module never invents a value that
isn't literally present in a cell.
"""
from __future__ import annotations

import re
from typing import Any, Optional

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from rapidfuzz import fuzz, process


def load_workbook(path: str):
    return openpyxl.load_workbook(path, data_only=True, read_only=True)


def _norm(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def find_sheet(wb, *candidates: str, threshold: int = 80) -> Optional[Worksheet]:
    """Fuzzy-match a sheet name against `candidates`, tried in order."""
    names = wb.sheetnames
    norm_names = [_norm(n).lower() for n in names]
    for cand in candidates:
        norm_cand = _norm(cand).lower()
        if norm_cand in norm_names:
            return wb[names[norm_names.index(norm_cand)]]
        match = process.extractOne(norm_cand, norm_names, scorer=fuzz.token_sort_ratio)
        if match and match[1] >= threshold:
            return wb[names[norm_names.index(match[0])]]
    return None


def sheet_matrix(ws: Worksheet, max_rows: Optional[int] = None) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        rows.append(list(row))
        if max_rows and i + 1 >= max_rows:
            break
    return rows


def find_header_row(
    matrix: list[list[Any]],
    expected_anchors: list[str],
    max_scan_rows: int = 15,
    threshold: int = 75,
) -> Optional[int]:
    """Best-guess 0-based index of the header row, by fuzzy-matching cell text
    in each of the first `max_scan_rows` rows against `expected_anchors`
    (known column names for this sheet type). Handles title rows / merged
    banner rows sitting above the real header, common in these workbooks."""
    best_idx, best_score = None, 0
    for i, row in enumerate(matrix[:max_scan_rows]):
        cells = [_norm(c) for c in row if _norm(c)]
        if not cells:
            continue
        lowered = [c.lower() for c in cells]
        score = 0
        for anchor in expected_anchors:
            match = process.extractOne(anchor.lower(), lowered, scorer=fuzz.token_sort_ratio)
            if match and match[1] >= threshold:
                score += 1
        if score > best_score:
            best_score, best_idx = score, i
    return best_idx


def resolve_columns(headers: list[Any], candidate_names: list[str], threshold: int = 75) -> dict[str, Optional[int]]:
    """Map each of `candidate_names` to the best-matching column index in
    `headers`, or None if nothing crosses `threshold`."""
    norm_headers = [_norm(h).lower() for h in headers]
    result: dict[str, Optional[int]] = {}
    for cand in candidate_names:
        cand_l = cand.lower()
        if cand_l in norm_headers:
            result[cand] = norm_headers.index(cand_l)
            continue
        match = process.extractOne(cand_l, norm_headers, scorer=fuzz.token_sort_ratio)
        result[cand] = norm_headers.index(match[0]) if match and match[1] >= threshold else None
    return result


_FEATURE_ID_RE = re.compile(r"(\d{1,3})")


def normalize_feature_id(value: Any) -> Optional[str]:
    """Normalize a feature-id-ish value ('005', '5', ' Feature_005 ', 5) to a
    zero-padded 3-digit string, or None if it doesn't contain digits."""
    s = _norm(value)
    if not s:
        return None
    match = _FEATURE_ID_RE.search(s)
    if not match:
        return None
    return match.group(1).zfill(3)


def rows_as_dicts(
    matrix: list[list[Any]], header_row_idx: int, col_map: dict[str, Optional[int]]
) -> list[dict[str, Any]]:
    """Convert every non-blank data row below `header_row_idx` into a dict
    keyed by `col_map`'s keys, reading only real cell values."""
    out: list[dict[str, Any]] = []
    for row in matrix[header_row_idx + 1 :]:
        if all(_norm(c) == "" for c in row):
            continue
        record = {key: (row[idx] if idx is not None and idx < len(row) else None) for key, idx in col_map.items()}
        out.append(record)
    return out


def _fuzzy_key(value: Any) -> str:
    """Normalized form used only for fuzzy scoring: underscores treated as
    word separators too, so 'Config_Tol_Spd' and 'Config Tol Spd' score as
    near-identical regardless of which separator style was typed/read."""
    return _norm(value).lower().replace("_", " ")


def fuzzy_find(needle: Any, haystack: list[str], threshold: int = 90) -> Optional[str]:
    """Return the entry in `haystack` that best fuzzy-matches `needle`, or None
    if nothing crosses `threshold`. Used by requirements_extract's Category
    classification - a value that doesn't cross the threshold against the
    known vocabulary is dropped rather than guessed."""
    n = _fuzzy_key(needle)
    if not n:
        return None
    keyed = [_fuzzy_key(h) for h in haystack]
    if n in keyed:
        return haystack[keyed.index(n)]
    match = process.extractOne(n, keyed, scorer=fuzz.token_sort_ratio)
    return haystack[keyed.index(match[0])] if match and match[1] >= threshold else None
